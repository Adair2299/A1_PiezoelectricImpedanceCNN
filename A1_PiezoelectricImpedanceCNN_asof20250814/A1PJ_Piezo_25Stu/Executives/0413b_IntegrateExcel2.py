import os
from openpyxl import load_workbook

# 定义文件路径
base_path = r"E:\01我的\大三下(202501-202508)\大创-压电阻抗\14天阻抗汇总\Term2"
main_file = os.path.join(base_path, "Term4 1-500 01.19.xlsx")
output_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
               'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
               'AA', 'AB', 'AC', 'AD']

# 加载主工作簿
wb_main = load_workbook(main_file)
ws_main = wb_main.active

# 遍历从01到30的文件
for i in range(1, 31):
    # 如果 i == 6，直接跳出循环（直接忽略6号钢板）
    if i == 6:
        continue  # 终止循环，不再处理后续文件

    # 构建文件名，确保两位数格式
    file_num = f"{i:02d}"
    source_file = os.path.join(base_path, f"1-500 01.19_{file_num}.xlsx")

    try:
        # 加载源工作簿
        wb_source = load_workbook(source_file)
        ws_source = wb_source.active

        # 获取B2单元格的值
        cell_value = ws_source['B2'].value

        # 写入主工作簿的第一行对应列
        ws_main[f"{output_cols[i - 1]}1"] = cell_value

        print(f"成功处理文件 {file_num}: 值 {cell_value} 已写入 {output_cols[i - 1]}1")

    except Exception as e:
        print(f"处理文件 {file_num} 时出错: {str(e)}")

# 保存主工作簿
wb_main.save(main_file)
print("所有文件处理完成，主文件已保存。")