import os
import openpyxl
from openpyxl.utils import get_column_letter

for pnum in list(range(1, 18)) + [27, 28, 30]:
    # 输入目录（所有Excel文件所在位置）
    input_dir = rf"E:\01我的\大三下(202501-202508)\大创-压电阻抗\有限元\PZT FEM2\{pnum:02d}"
    # 输出文件（新Excel文件）
    output_file = os.path.join(input_dir, f"整合数据_按列排列{pnum:02}.xlsx")

    # 创建新的工作簿
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "整合数据"

    # 当前列（从第1列开始）
    current_col = 1

    # 遍历目录中的所有Excel文件
    for filename in os.listdir(input_dir):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            if filename == os.path.basename(output_file):  # 跳过输出文件
                continue

            filepath = os.path.join(input_dir, filename)
            print(f"正在处理文件: {filename}")

            try:
                # 读取Excel文件
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active  # 假设数据在第一个工作表

                # 写入文件名到第一行（作为列标题）
                output_ws.cell(row=1, column=current_col, value=filename)

                # 读取B9到B2504的数据，并写入当前列
                for row in range(9, 2505):
                    cell_value = ws[f"B{row}"].value
                    output_ws.cell(row=row - 7, column=current_col, value=cell_value)  # row-7 是为了让数据从第2行开始

                current_col += 1  # 处理下一个文件时换列

            except Exception as e:
                print(f"处理文件 {filename} 时出错: {e}")

    # 保存整合后的Excel文件
    output_wb.save(output_file)
    print(f"数据已按列整合保存到: {output_file}")